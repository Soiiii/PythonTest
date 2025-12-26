import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

input_file = "input.xlsx"
output_file = "output.xlsx"

# 1. 엑셀 import
df = pd.read_excel(input_file)

# 2. 그대로 export (서식 작업은 openpyxl)
df.to_excel(output_file, index=False)

# 3. 엑셀 다시 열기
wb = load_workbook(output_file)
ws = wb.active

yellow_fill = PatternFill(
    start_color="FFFF00",
    end_color="FFFF00",
    fill_type="solid"
)

def normalize(text):
    return (
        str(text)
        .strip()
        .lower()
        .replace("-", " ")
    )


def split_answers(text):
    return [normalize(t) for t in str(text).split("/") if t.strip()]


def is_correct(correct_answer, user_input):
    """
    correct_answer: B열 (정답, 복수 가능)
    user_input: A열 또는 C열
    """
    if user_input is None or str(user_input).strip() == "":
        return False

    correct_parts = split_answers(correct_answer)
    user_parts = split_answers(user_input)

    # 하나라도 겹치면 정답
    return any(u in correct_parts for u in user_parts)

# 4. 행별 채점
for row in range(2, ws.max_row + 1):
    user_a = ws.cell(row=row, column=1).value   # A열 (답안)
    correct_b = ws.cell(row=row, column=2).value  # B열 (정답)
    user_c = ws.cell(row=row, column=3).value   # C열 (답안)

    # A vs B
    if not is_correct(correct_b, user_a):
        ws.cell(row=row, column=1).fill = yellow_fill

    # C vs B
    if not is_correct(correct_b, user_c):
        ws.cell(row=row, column=3).fill = yellow_fill


# 5. 저장
wb.save(output_file)

print("채점 완료! B열 기준으로 A/C 개별 채점됨")
